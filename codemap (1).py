
import os
import re
import json
import time
import csv
from datetime import datetime
from collections import deque, defaultdict
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from statistics import mean
import shutil
import warnings
warnings.filterwarnings('ignore')

# ========== CONFIG ==========
BASE_DIR = r"C:\Users\Public\EVCS_Multi_Station"
os.makedirs(BASE_DIR, exist_ok=True)

# 3 lớp backup
OUTPUT_EXCEL = os.path.join(BASE_DIR, "evcs_danang_realtime.xlsx")
STATE_FILE = os.path.join(BASE_DIR, "stations_state.json")
BACKUP_DIR = os.path.join(BASE_DIR, "backups")
STATS_CSV = os.path.join(BASE_DIR, "charging_stats.csv")

os.makedirs(BACKUP_DIR, exist_ok=True)

UPDATE_INTERVAL_MINUTES = 15

STATIONS = [
    ("Trạm 1", "https://evcs.vn/tram-sac-vinfast-ho-kinh-doanh-tram-sac-vivu-74-dang-huy-tru-p-hoa-khanh-da-nang-c.dna0181.html"),
    ("Trạm 2", "https://evcs.vn/tram-sac-vinfast-bai-do-xe-tu-nhan-hoa-bac-c.dna0054.html"),
    ("Trạm 3", "https://evcs.vn/tram-sac-vinfast-nhuong-quyen-cty-cao-su-da-nang-c.dna0103.html"),
    ("Trạm 4", "https://evcs.vn/tram-sac-vinfast-nhuong-quyen-le-thanh-hung-c.dna0113.html"),
    ("Trạm 5", "https://evcs.vn/tram-sac-vinfast-cua-hang-xang-dau-hoa-khanh-so-15-c.dna0016.html"),
    ("Trạm 6", "https://evcs.vn/tram-sac-vinfast-bdx-tn-ba-na-c.dna0109.html"),
    ("Trạm 7", "https://evcs.vn/tram-sac-vinfast-nhuong-quyen-do-huu-tuan-c.dna0138.html"),
    ("Trạm 8", "https://evcs.vn/tram-sac-vinfast-nhuong-quyen-goevtruong-anh-tri-c.dna0141.html"),
    ("Trạm 9", "https://evcs.vn/tram-sac-vinfast-bai-do-xe-nut-giao-tuy-loan-c.dna0115.html"),
    ("Trạm 10", "https://evcs.vn/tram-sac-vinfast-cua-hang-xang-dau-pvoil-mt-pham-hung-c.dna0045.html"),
    ("Trạm 11", "https://evcs.vn/tram-sac-vinfast-cua-hang-xang-dau-pvoil-nqtm-hoa-hiep-2-c.dna0048.html"),
    ("Trạm 12", "https://evcs.vn/tram-sac-vinfast-nhuong-quyen-nguyen-doan-huy-phuong-c.dna0170.html"),
    ("Trạm 13", "https://evcs.vn/tram-sac-vinfast-cua-hang-xang-dau-pvoil-hoa-chau-c.dna0025.html"),
    ("Trạm 14", "https://evcs.vn/tram-sac-vinfast-nhuong-quyen-dimec-viet-huong-da-nang-1-c.dna0132.html"),
    ("Trạm 15", "https://evcs.vn/tram-sac-vinfast-nq-dimec-izu-viet-nam-1-c.dna0143.html"),
    ("Trạm 16", "https://evcs.vn/tram-sac-vinfast-cua-hang-xang-dau-petrolimex-so-01-khu-vuc-5-c.dna0038.html"),
    ("Trạm 17", "https://evcs.vn/tram-sac-vinfast-tu-nhan-ngo-gia-bao-14-hoa-an-3cam-le-da-nang-c.dna3876.html"),
    ("Trạm 18", "https://evcs.vn/tram-sac-vinfast-nhuong-quyen-vip-cong-ty-dainco-c.dna0057.html"),
    ("Trạm 19", "https://evcs.vn/tram-sac-vinfast-nq-tram-dang-kiem-4305d-c.dna0111.html"),
    ("Trạm 20", "https://evcs.vn/tram-sac-vinfast-tu-nhan-nguyen-khac-an-94-nguyen-thi-can-phuong-hoa-an-quan-cam-le-c.dna0197.html"),
    ("Trạm 21", "https://evcs.vn/tram-sac-vinfast-nhuong-quyen-fast-nguyen-quang-vinh-c.dna0187.html"),
    ("Trạm 22", "https://evcs.vn/tram-sac-vinfast-nq-chxd-van-xuan-c.dna0096.html"),
    ("Trạm 23", "https://evcs.vn/tram-sac-vinfast-nhuong-quyen-txtx-nguyen-hiep-duc-c.dna0064.html"),
    ("Trạm 24", "https://evcs.vn/tram-sac-vinfast-nhuong-quyen-luxury-cong-ty-minh-thanh-toan-c.dna0061.html"),
    ("Trạm 25", "https://evcs.vn/tram-sac-vinfast-nq-cua-hang-xang-dau-hoa-cam-c.dna0173.html")
]


class SafeBackupSystem:
    """Hệ thống backup 3 lớp - KHÔNG BAO GIỜ MẤT DỮ LIỆU"""
    
    def __init__(self):
        self.backup_history = []
        self.load_backup_history()
    
    def save_state_safe(self, state_data):
        """Lưu state với 3 lớp backup"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # LỚP 1: Lưu JSON chính
        try:
            with open(STATE_FILE, "w", encoding="utf-8") as f:
                json.dump(state_data, f, ensure_ascii=False, indent=2)
            print(f"✅ Lớp 1: Lưu JSON chính")
        except Exception as e:
            print(f"❌ Lỗi lưu JSON chính: {e}")
        
        # LỚP 2: Lưu backup theo timestamp
        backup_file = os.path.join(BACKUP_DIR, f"state_backup_{timestamp}.json")
        try:
            with open(backup_file, "w", encoding="utf-8") as f:
                json.dump(state_data, f, ensure_ascii=False, indent=2)
            self.backup_history.append(backup_file)
            print(f"✅ Lớp 2: Backup timestamped")
            
            # Giữ tối đa 50 backup gần nhất
            if len(self.backup_history) > 50:
                old_file = self.backup_history.pop(0)
                if os.path.exists(old_file):
                    os.remove(old_file)
        except Exception as e:
            print(f"❌ Lỗi backup timestamp: {e}")
        
        # LỚP 3: Lưu CSV dễ đọc (fallback)
        self.save_to_csv(state_data)
        
        # Lưu lịch sử backup
        self.save_backup_history()
    
    def save_to_csv(self, state_data):
        """LỚP 3: Lưu vào CSV (backup dễ đọc)"""
        try:
            with open(STATS_CSV, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow(["Tên trạm", "Loại cổng", "Tổng phút", "Số xe", "TB (phút)", "Cập nhật"])
                
                for station_name, station_data in state_data.get("stations", {}).items():
                    avg_data = station_data.get("avg_data", {})
                    for charger_type, data in avg_data.items():
                        total_min = data.get("total_minutes", 0)
                        count = data.get("count", 0)
                        avg = round(total_min / count, 2) if count > 0 else 0
                        
                        writer.writerow([
                            station_name,
                            charger_type,
                            round(total_min, 2),
                            count,
                            avg,
                            state_data.get("last_update", "N/A")
                        ])
            print(f"✅ Lớp 3: Lưu CSV backup")
        except Exception as e:
            print(f"❌ Lỗi lưu CSV: {e}")
    
    def load_state_safe(self):
        """Khôi phục state từ 3 lớp backup (ưu tiên cao → thấp)"""
        
        # Thử LỚP 1: JSON chính
        if os.path.exists(STATE_FILE):
            try:
                with open(STATE_FILE, "r", encoding="utf-8") as f:
                    state = json.load(f)
                print(f"✅ Khôi phục từ JSON chính: {state.get('last_update', 'N/A')}")
                return state
            except Exception as e:
                print(f"⚠️ JSON chính lỗi: {e}")
        
        # Thử LỚP 2: Backup gần nhất
        if self.backup_history:
            for backup_file in reversed(self.backup_history):
                if os.path.exists(backup_file):
                    try:
                        with open(backup_file, "r", encoding="utf-8") as f:
                            state = json.load(f)
                        print(f"✅ Khôi phục từ backup: {os.path.basename(backup_file)}")
                        return state
                    except Exception as e:
                        print(f"⚠️ Backup {backup_file} lỗi: {e}")
                        continue
        
        # Thử LỚP 3: CSV backup
        if os.path.exists(STATS_CSV):
            try:
                state = self.load_from_csv()
                print(f"✅ Khôi phục từ CSV backup")
                return state
            except Exception as e:
                print(f"⚠️ CSV backup lỗi: {e}")
        
        print("⚠️ Không tìm thấy backup nào - Bắt đầu từ đầu")
        return None
    
    def load_from_csv(self):
        """Khôi phục từ CSV (lớp 3)"""
        state = {
            "last_update": "Khôi phục từ CSV",
            "stations": defaultdict(lambda: {"avg_data": defaultdict(dict)})
        }
        
        with open(STATS_CSV, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                station_name = row["Tên trạm"]
                charger_type = row["Loại cổng"]
                
                state["stations"][station_name]["avg_data"][charger_type] = {
                    "total_minutes": float(row["Tổng phút"]),
                    "count": int(row["Số xe"])
                }
        
        return dict(state)
    
    def save_backup_history(self):
        """Lưu danh sách các backup"""
        history_file = os.path.join(BACKUP_DIR, "backup_history.json")
        try:
            with open(history_file, "w", encoding="utf-8") as f:
                json.dump({"backups": self.backup_history}, f, indent=2)
        except:
            pass
    
    def load_backup_history(self):
        """Load danh sách các backup"""
        history_file = os.path.join(BACKUP_DIR, "backup_history.json")
        if os.path.exists(history_file):
            try:
                with open(history_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    self.backup_history = data.get("backups", [])
            except:
                pass


class StationMonitor:
    """Giám sát trạm với FIFO + backup an toàn"""
    
    def __init__(self, station_name, url):
        self.station_name = station_name
        self.url = url
        self.prev_data = {}
        self.queues = defaultdict(deque)
        self.avg_data = defaultdict(lambda: {"total_minutes": 0.0, "count": 0})
        self.power_map = {}
    
    def parse_webpage(self, html):
        """Parse HTML"""
        soup = BeautifulSoup(html, "html.parser")
        text = soup.get_text(" ", strip=True)
        
        pattern = r'(\d+(?:\.\d+)?)\s*kW\s*trống\s*(\d+)\s*/\s*(\d+)'
        
        result = {}
        for match in re.finditer(pattern, text, re.IGNORECASE):
            power = match.group(1)
            free = int(match.group(2))
            total = int(match.group(3))
            
            charger_type = f"{power}kW"
            result[charger_type] = {
                "free": free,
                "total": total,
                "charging": total - free,
                "power": float(power)
            }
            
            if charger_type not in self.power_map:
                self.power_map[charger_type] = float(power)
        
        return result
    
    def check_and_update(self, timestamp):
        """Kiểm tra và cập nhật"""
        try:
            resp = requests.get(self.url, timeout=15)
            current_data = self.parse_webpage(resp.text)
            
            if not current_data:
                return None
            
            result = {
                "timestamp": timestamp,
                "station": self.station_name,
                "total_power": 0,
                "chargers": {},
                "avg_durations": {},
                "total_durations": {},
                "count_vehicles": {}
            }
            
            # Tính tổng công suất
            for ctype, data in current_data.items():
                charging = data.get("charging", 0)
                power = data.get("power", 0)
                result["total_power"] += charging * power
            
            # Xử lý từng loại sạc
            if self.prev_data:
                for ctype in current_data.keys():
                    self._process_charger_type(ctype, current_data, timestamp)
            
            # Lưu dữ liệu hiện tại
            for ctype, data in current_data.items():
                result["chargers"][ctype] = {
                    "free": data["free"],
                    "total": data["total"],
                    "power": data["power"]
                }
                
                # Thống kê chi tiết
                avg_info = self.avg_data[ctype]
                result["total_durations"][ctype] = round(avg_info["total_minutes"], 2)
                result["count_vehicles"][ctype] = avg_info["count"]
                
                if avg_info["count"] > 0:
                    result["avg_durations"][ctype] = round(avg_info["total_minutes"] / avg_info["count"], 2)
                else:
                    result["avg_durations"][ctype] = 0
            
            self.prev_data = current_data
            return result
            
        except Exception as e:
            print(f"❌ {self.station_name}: {str(e)[:50]}")
            return None
    
    def _process_charger_type(self, charger_type, current_data, timestamp):
        """FIFO processing"""
        prev = self.prev_data.get(charger_type, {"free": 0, "total": 0})
        curr = current_data.get(charger_type, {"free": 0, "total": 0})
        
        prev_free = prev["free"]
        curr_free = curr["free"]
        
        # Xe cắm vào
        if curr_free < prev_free:
            plug_count = prev_free - curr_free
            for i in range(plug_count):
                self.queues[charger_type].append(timestamp)
        
        # Xe rút ra
        elif curr_free > prev_free:
            unplug_count = curr_free - prev_free
            for i in range(unplug_count):
                if self.queues[charger_type]:
                    start_time = self.queues[charger_type].popleft()
                    
                    st = datetime.strptime(start_time, "%Y-%m-%d %H:%M:%S")
                    et = datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S")
                    duration = (et - st).total_seconds() / 60.0
                    
                    # CỘNG DỒN - QUAN TRỌNG!
                    self.avg_data[charger_type]["total_minutes"] += duration
                    self.avg_data[charger_type]["count"] += 1
                    
                    print(f"  📝 {self.station_name} - {charger_type}: "
                          f"Xe sạc {duration:.1f}p → "
                          f"Tổng: {self.avg_data[charger_type]['total_minutes']:.1f}p / "
                          f"{self.avg_data[charger_type]['count']} xe = "
                          f"{self.avg_data[charger_type]['total_minutes']/self.avg_data[charger_type]['count']:.1f}p TB")
    
    def get_state(self):
        """Lấy trạng thái ĐẦY ĐỦ"""
        return {
            "prev_data": self.prev_data,
            "queues": {k: list(v) for k, v in self.queues.items()},
            "avg_data": {k: dict(v) for k, v in self.avg_data.items()},  # Convert defaultdict
            "power_map": self.power_map
        }
    
    def load_state(self, state):
        """Khôi phục trạng thái ĐẦY ĐỦ"""
        self.prev_data = state.get("prev_data", {})
        self.power_map = state.get("power_map", {})
        
        # Load queues
        queues_data = state.get("queues", {})
        for ctype, queue_list in queues_data.items():
            self.queues[ctype] = deque(queue_list)
        
        # Load avg_data - QUAN TRỌNG!
        avg_data = state.get("avg_data", {})
        for ctype, data in avg_data.items():
            self.avg_data[ctype] = {
                "total_minutes": float(data.get("total_minutes", 0)),
                "count": int(data.get("count", 0))
            }


class MultiStationManager:
    """Quản lý với backup system"""
    
    def __init__(self):
        self.backup_system = SafeBackupSystem()
        self.stations = []
        self.all_charger_types = set()
        
        for name, url in STATIONS:
            monitor = StationMonitor(name, url)
            self.stations.append(monitor)
        
        self.load_state()
        print(f"✅ Đã tải {len(self.stations)} trạm sạc")
    
    def scan_all(self):
        """Quét tất cả"""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"\n{'='*70}")
        print(f"🔍 Quét {len(self.stations)} trạm tại {timestamp}")
        print(f"{'='*70}")
        
        results = []
        new_charger_types = set()
        
        for i, station in enumerate(self.stations, 1):
            print(f"[{i:3d}/{len(self.stations)}] {station.station_name[:25]:<25}", end=" ")
            result = station.check_and_update(timestamp)
            if result:
                results.append(result)
                new_charger_types.update(result["chargers"].keys())
                charger_info = ", ".join([f"{k}: {v['free']}/{v['total']}" for k, v in result["chargers"].items()])
                print(f"✓ [{charger_info}]")
            else:
                print("✗")
            time.sleep(0.3)
        
        self.all_charger_types.update(new_charger_types)
        
        if results:
            self.save_to_excel(results, timestamp)
            self.save_state()
            self.print_summary()
        
        print(f"{'='*70}\n")
    
    def save_to_excel(self, results, timestamp):
        """Lưu Excel"""
        if os.path.exists(OUTPUT_EXCEL):
            wb = openpyxl.load_workbook(OUTPUT_EXCEL)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "EVCS Real-time"
            
            headers = ["Thời gian", "Tên trạm", "Tổng công suất (kW)"]
            
            sorted_chargers = sorted(self.all_charger_types, 
                                    key=lambda x: float(x.replace('kW', '')), 
                                    reverse=True)
            
            for ctype in sorted_chargers:
                headers.extend([
                    f"{ctype} Trống/Tổng",
                    f"{ctype} TB (phút)"
                ])
            
            ws.append(headers)
            
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        sorted_chargers = sorted(self.all_charger_types, 
                                key=lambda x: float(x.replace('kW', '')), 
                                reverse=True)
        
        for result in results:
            row = [
                timestamp,
                result["station"],
                round(result["total_power"], 2)
            ]
            
            for ctype in sorted_chargers:
                charger = result["chargers"].get(ctype, {"free": 0, "total": 0})
                if charger["total"] == 0:
                    row.append("0/0")
                else:
                    row.append(f"{charger['free']}/{charger['total']}")
                
                avg_duration = result["avg_durations"].get(ctype, 0)
                row.append(avg_duration if avg_duration > 0 else 0)
            
            ws.append(row)
        
        # Auto-size
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column].width = adjusted_width
        
        wb.save(OUTPUT_EXCEL)
        print(f"💾 Đã lưu Excel: {OUTPUT_EXCEL}")
    
    def save_state(self):
        """Lưu state với backup 3 lớp"""
        state = {
            "last_update": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "stations": {s.station_name: s.get_state() for s in self.stations}
        }
        
        print(f"\n💾 Đang backup dữ liệu...")
        self.backup_system.save_state_safe(state)
        print(f"✅ Backup hoàn tất!\n")
    
    def load_state(self):
        """Khôi phục state"""
        state = self.backup_system.load_state_safe()
        
        if state:
            for station in self.stations:
                if station.station_name in state.get("stations", {}):
                    station.load_state(state["stations"][station.station_name])
                    self.all_charger_types.update(station.power_map.keys())
            
            print(f"✅ Đã khôi phục từ: {state.get('last_update', 'N/A')}")
            
            # In thống kê đã khôi phục
            total_vehicles = 0
            for station in self.stations:
                for ctype, data in station.avg_data.items():
                    total_vehicles += data["count"]
            
            if total_vehicles > 0:
                print(f"📊 Dữ liệu đã có: {total_vehicles} lượt xe sạc")
    
    def print_summary(self):
        """In tóm tắt"""
        print(f"\n📊 TÓM TẮT THỐNG KÊ")
        print(f"{'─'*70}")
        
        sorted_types = sorted(self.all_charger_types, 
                            key=lambda x: float(x.replace('kW', '')), 
                            reverse=True)
        
        for ctype in sorted_types:
            total_minutes = 0
            total_count = 0
            
            for station in self.stations:
                if ctype in station.avg_data:
                    total_minutes += station.avg_data[ctype]["total_minutes"]
                    total_count += station.avg_data[ctype]["count"]
            
            if total_count > 0:
                avg = total_minutes / total_count
                print(f"  {ctype:8} │ TB: {avg:5.1f}p │ Tổng: {total_minutes:8.1f}p │ Số xe: {total_count:4d}")


def main():
    """Hàm chính"""
    import argparse
    
    parser = argparse.ArgumentParser(description="EVCS Monitor - Safe Backup System")
    parser.add_argument("--once", action="store_true", help="Chạy 1 lần")
    parser.add_argument("--interval", type=int, default=UPDATE_INTERVAL_MINUTES, 
                       help="Khoảng cách (phút)")
    parser.add_argument("--restore", action="store_true", help="Xem thông tin backup")
    args = parser.parse_args()
    
    if args.restore:
        backup_system = SafeBackupSystem()
        state = backup_system.load_state_safe()
        if state:
            print("\n📋 THÔNG TIN BACKUP")
            print(f"Cập nhật lần cuối: {state.get('last_update', 'N/A')}")
            print(f"Số trạm: {len(state.get('stations', {}))}")
            
            total_vehicles = 0
            for station_name, station_data in state.get("stations", {}).items():
                for ctype, data in station_data.get("avg_data", {}).items():
                    total_vehicles += data.get("count", 0)
            print(f"Tổng số xe đã ghi nhận: {total_vehicles}")
        return
    
    manager = MultiStationManager()
    
    try:
        if args.once:
            manager.scan_all()
        else:
            print(f"🚀 EVCS Monitor - Hệ thống backup 3 lớp")
            print(f"📍 {len(manager.stations)} trạm sạc VinFast Đà Nẵng")
            print(f"⏰ Cập nhật mỗi {args.interval} phút")
            print(f"💾 Backup: JSON + Timestamp + CSV")
            print(f"🛡️ KHÔNG BAO GIỜ MẤT DỮ LIỆU!")
            print("Nhấn Ctrl+C để dừng\n")
            
            while True:
                manager.scan_all()
                
                wait_seconds = args.interval * 60
                next_time = datetime.now()
                next_time = datetime.fromtimestamp(next_time.timestamp() + wait_seconds)
                
                print(f"⏳ Chờ đến {next_time.strftime('%H:%M:%S')} ({args.interval} phút)...")
                time.sleep(wait_seconds)
                
    except KeyboardInterrupt:
        print("\n🛑 Dừng giám sát!")
        print("💾 Đang lưu backup cuối cùng...")
        manager.save_state()
        print("✅ Hoàn tất! Dữ liệu đã được lưu an toàn.")


if __name__ == "__main__":

    main()
